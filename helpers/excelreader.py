import openpyxl


class ExcelReader:
    """
    Excel extraction point where the requirement text collected
    """
    def __init__(self, loc: str):
        self.loc = loc
        self.workbook = openpyxl.load_workbook(self.loc)
        self.first_sheet = self.workbook.worksheets[0]

    def extract_data_by_rowkey(self, row_key: str) -> dict:
        """
        Data extraction from row_key header. Needed specifically to
        retrieve back a dict with the requirement steps

        :param row_key: str for the desired header under where the data is located
        :return: dictionary with collected data from header "row_key""
        """
        collection_of_row_keys = {}
        req_count = 0

        # Loop through columns in the first row
        for col in range(1, self.first_sheet.max_column + 1):
            cell_value = self.first_sheet.cell(row=1, column=col).value
            if cell_value == row_key:
                # Start from row 2 to skip header
                for row in range(2, self.first_sheet.max_row + 1):
                    value = self.first_sheet.cell(row=row, column=col).value

                    #skip completely empty rows
                    if value is None:
                        continue
                    collection_of_row_keys[f"Requirement {req_count}"] = value
                    req_count += 1
                break

        return collection_of_row_keys

    def remove_duplicates(self) -> str:
        """
        :return: removes duplicates from dict and translates it to return str
        """
        seen_values = []
        new_dict = {}
        found_data = self.extract_data_by_rowkey('Requirement Description')
        for key, value in found_data.items():
            if value not in seen_values:
                seen_values.append(value)
                new_dict[key] = value

        return ", ".join(f"{k}: {v}" for k, v in new_dict.items())

#
# if __name__ == '__main__':
#     excel_obj = Excel_Reader(r'C:\Users\raduicar\OneDrive - Bertrandt AG\Desktop\dummy_sample.xlsx')
#     removed_duplicates = excel_obj.remove_duplicates()
